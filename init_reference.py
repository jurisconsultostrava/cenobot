#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Jednorázová inicializace referenčních cen (last_price) do products_db.json.
Slouží k tomu, aby sanity-check (MAX_ZMENA_PCT) fungoval hned od prvního běhu.
Vstup: aktuální Shoptet export (products__19_.xlsx nebo přes API).
Bez tohoto kroku první běh sanity-check přeskočí (nemá s čím srovnávat).
"""
import json, sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
DB = BASE / "products_db.json"

def from_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    hdr = next(ws.iter_rows(values_only=True))
    ci = {h: i for i, h in enumerate(hdr)}
    prices = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = str(r[ci["code"]]).strip() if r[ci.get("code", -1)] else ""
        if not code:
            continue
        try:
            prices[code] = float(str(r[ci["price"]]).replace(",", "."))
        except (TypeError, ValueError, KeyError):
            pass
    wb.close()
    return prices

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Použití: python3 init_reference.py <cesta_k_exportu.xlsx>")
        sys.exit(1)
    db = json.loads(DB.read_text(encoding="utf-8"))
    prices = from_xlsx(sys.argv[1])
    n = 0
    for code, p in prices.items():
        if code in db and p and p > 0:
            db[code]["last_price"] = p
            n += 1
    DB.write_text(json.dumps(db, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"Referenční ceny nastaveny u {n} produktů.")
